import pandas as pd
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Add project root
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from python.utils.db_connection import get_db_connection, get_cursor

OUTPUT_DIR = Path("docker/data/templates")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TABLES = {
    'base_oficial': 'bronze.base_oficial',
    'faturamento': 'bronze.faturamento',
    'usuarios': 'bronze.usuarios'
}

def generate_templates():
    conn = get_db_connection()
    
    for name, table in TABLES.items():
        print(f"Gerando template para: {name}...")
        
        with get_cursor(conn) as cur:
            cur.execute(f"SELECT * FROM {table} LIMIT 0")
            cols = [desc[0] for desc in cur.description 
                   if desc[0] not in ('id', 'data_carga', 'source_filename')]
        
        # Create Excel
        wb = Workbook()
        ws = wb.active
        ws.title = name
        
        # Headers
        ws.append(cols)
        
        # Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="003366")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        # Save
        path = OUTPUT_DIR / f"template_{name}.xlsx"
        wb.save(path)
        print(f" -> Salvo: {path}")

    conn.close()

if __name__ == "__main__":
    generate_templates()
